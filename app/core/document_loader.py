import os
import re

import numpy as np
from loguru import logger


class DocumentLoader:
    """Load and chunk documents with embedding-based semantic boundary detection.

    Chunking strategy:
      1. Split text into sentences
      2. Compute sentence embeddings
      3. Detect topic boundaries via cosine similarity drop between adjacent groups
      4. Split at boundaries, respecting max chunk size
      5. Add overlap for context continuity
    """

    CHUNK_SIZE = 800          # target characters per chunk
    CHUNK_OVERLAP = 150       # overlap between chunks
    MIN_CHUNK_SIZE = 100      # discard chunks smaller than this
    SIMILARITY_PERCENTILE = 25   # breakpoint at bottom N% of similarity scores
    WINDOW_SIZE = 3           # sentences per group for smoothing

    def load_and_chunk(self, file_path: str, filename: str) -> list[str]:
        ext = os.path.splitext(filename)[1].lower()

        if ext == ".pdf":
            text = self._load_pdf(file_path)
        elif ext == ".docx":
            text = self._load_docx(file_path)
        elif ext in (".xlsx", ".xls"):
            text = self._load_excel(file_path)
        else:
            text = self._load_text(file_path)

        chunks = self._semantic_chunk(text)
        logger.info(f"Loaded {filename}: {len(text):,} chars → {len(chunks)} semantic chunks")
        return chunks

    # ── Loaders ─────────────────────────────────────────────

    def _load_pdf(self, file_path: str) -> str:
        from pypdf import PdfReader
        reader = PdfReader(file_path)
        pages = []
        for i, page in enumerate(reader.pages):
            text = page.extract_text() or ""
            if text.strip():
                pages.append(f"[Page {i+1}]\n{text}")
        return "\n\n".join(pages)

    def _load_docx(self, file_path: str) -> str:
        from docx import Document
        doc = Document(file_path)
        parts = []
        for para in doc.paragraphs:
            if para.text.strip():
                parts.append(para.text)
        return "\n\n".join(parts)

    def _load_excel(self, file_path: str) -> str:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True)
        lines = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            lines.append(f"=== Sheet: {sheet} ===")
            for row in ws.iter_rows(values_only=True):
                line = "\t".join(str(c) if c is not None else "" for c in row)
                if line.strip():
                    lines.append(line)
        return "\n".join(lines)

    def _load_text(self, file_path: str) -> str:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()

    # ── Semantic Chunking (Embedding-Based Boundary Detection) ──

    def _split_sentences(self, text: str) -> list[str]:
        """Split text into sentences (Korean + English aware)."""
        parts = re.split(r'(?<=[.!?。])\s+|(?<=\n)\n+', text)
        sentences = [p.strip() for p in parts if p.strip()]
        return sentences if sentences else [text]

    def _get_embedding_model(self):
        """Lazy-load embedding model from VectorStore singleton."""
        try:
            from app.core.vector_store import get_vector_store
            vs = get_vector_store()
            return vs.st_model
        except Exception as e:
            logger.warning(f"Embedding model unavailable, falling back to size-based chunking: {e}")
            return None

    def _compute_breakpoints(self, sentences: list[str]) -> list[int]:
        """Detect topic boundaries by measuring cosine similarity between
        adjacent sentence groups.

        How it works:
          - Group every WINDOW_SIZE sentences into a "block" and compute its
            mean embedding. Grouping smooths out noise from single sentences.
          - Compute cosine similarity between consecutive block embeddings.
          - Where similarity drops below SIMILARITY_THRESHOLD, mark a boundary.

        Returns: sorted list of sentence indices where a topic shift occurs.
        """
        model = self._get_embedding_model()
        if model is None or len(sentences) < self.WINDOW_SIZE * 2:
            return []  # too few sentences or no model → no semantic boundaries

        # Build overlapping groups of WINDOW_SIZE sentences
        groups = []
        for i in range(len(sentences) - self.WINDOW_SIZE + 1):
            group_text = " ".join(sentences[i:i + self.WINDOW_SIZE])
            groups.append(group_text)

        # Encode all groups in one batch for efficiency
        embeddings = model.encode(groups, normalize_embeddings=True, show_progress_bar=False)

        # Cosine similarity between consecutive groups (already L2-normalized)
        similarities = []
        for i in range(len(embeddings) - 1):
            sim = float(np.dot(embeddings[i], embeddings[i + 1]))
            similarities.append(sim)

        # Adaptive threshold: use percentile instead of fixed value.
        # This adapts to each document's similarity distribution —
        # a technical manual has higher baseline similarity than mixed-topic docs.
        threshold = float(np.percentile(similarities, self.SIMILARITY_PERCENTILE))

        # Breakpoints: where similarity drops below adaptive threshold
        breakpoints = []
        for i, sim in enumerate(similarities):
            if sim < threshold:
                boundary_idx = i + self.WINDOW_SIZE
                if boundary_idx < len(sentences):
                    breakpoints.append(boundary_idx)

        logger.debug(
            f"Semantic breakpoints: {len(breakpoints)} found in {len(sentences)} sentences "
            f"(adaptive_threshold={threshold:.3f}, min_sim={min(similarities):.3f}, "
            f"avg_sim={np.mean(similarities):.3f})"
        )
        return sorted(set(breakpoints))

    def _merge_into_chunks(self, segments: list[list[str]]) -> list[str]:
        """Merge sentence segments into chunks respecting CHUNK_SIZE with overlap."""
        chunks: list[str] = []

        for segment in segments:
            # Each segment is a topically coherent group of sentences.
            # If a segment fits within CHUNK_SIZE, keep it as one chunk.
            segment_text = "\n".join(segment)
            if len(segment_text) <= self.CHUNK_SIZE:
                chunks.append(segment_text)
                continue

            # Oversized segment: split by size while respecting sentence boundaries
            current_parts: list[str] = []
            current_size = 0

            for sent in segment:
                sent_len = len(sent)
                if current_size + sent_len > self.CHUNK_SIZE and current_parts:
                    chunks.append("\n".join(current_parts))

                    # Overlap: carry over trailing sentences
                    overlap_parts: list[str] = []
                    overlap_size = 0
                    for part in reversed(current_parts):
                        if overlap_size + len(part) <= self.CHUNK_OVERLAP:
                            overlap_parts.insert(0, part)
                            overlap_size += len(part)
                        else:
                            break
                    current_parts = overlap_parts
                    current_size = overlap_size

                current_parts.append(sent)
                current_size += sent_len

            if current_parts:
                chunks.append("\n".join(current_parts))

        return chunks

    def _semantic_chunk(self, text: str) -> list[str]:
        """Embedding-based semantic chunking pipeline:

        1. Split into sentences
        2. Detect topic boundaries via embedding similarity
        3. Group sentences between boundaries into segments
        4. Merge segments into chunks (respecting CHUNK_SIZE + overlap)
        5. Filter out tiny chunks

        Fallback: if embedding model is unavailable, uses paragraph-based
        chunking with size limits (still better than naive fixed-size).
        """
        if not text.strip():
            return []

        # Split paragraphs first, then sentences within each
        paragraphs = re.split(r'\n{2,}', text.strip())
        all_sentences: list[str] = []
        for para in paragraphs:
            para = para.strip()
            if not para:
                continue
            if len(para) <= self.CHUNK_SIZE:
                all_sentences.append(para)
            else:
                all_sentences.extend(self._split_sentences(para))

        if not all_sentences:
            return []

        # Detect semantic breakpoints
        breakpoints = self._compute_breakpoints(all_sentences)

        # Split sentences into segments at breakpoints
        segments: list[list[str]] = []
        prev = 0
        for bp in breakpoints:
            if bp > prev:
                segments.append(all_sentences[prev:bp])
            prev = bp
        if prev < len(all_sentences):
            segments.append(all_sentences[prev:])

        # If no breakpoints found, treat all sentences as one segment
        if not segments:
            segments = [all_sentences]

        # Merge segments into properly sized chunks with overlap
        chunks = self._merge_into_chunks(segments)

        # Filter tiny chunks
        result = [c for c in chunks if len(c) >= self.MIN_CHUNK_SIZE]
        return result
